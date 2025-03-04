'''
uses py 3.9
'''

import sys, os
from PyQt5.QtWidgets import *
from PyQt5.QtCore import Qt, QTimer, QThread, pyqtSignal
from PyQt5.QtGui import QFont, QColor
sys.path.append(r'C:\Users\zix63\Documents\PyScripts')
import monkeyUnityv1_6 as mky
import ROIConfig
from datetime import datetime
import openpyxl
# import vid_play_v0_7 as vid_player # save for later.

DARK_STYLE = """
QWidget {
    background-color: #2D2D30;
    color: #DCDCDC;
}
QTabWidget::pane {
    border: 1px solid #3F3F46;
    background: #2D2D30;
}
QHeaderView::section {
        background-color: #3F3F46;
        color: white;
        border: 1px solid #2D2D30;
        padding: 4px;
}
QTableCornerButton::section {
    background-color: #3F3F46;
    border: 1px solid #2D2D30;
}
QTabBar::tab {
    background: #252526;
    color: #DCDCDC;
    padding: 8px;
    border: 1px solid #3F3F46;
}
QTabBar::tab:selected {
    background: #3E3E42;
}
QLineEdit {
    background: #252526;
    border: 1px solid #3F3F46;
    padding: 3px;
}
QPushButton {
    background: #007ACC;
    border: none;
    padding: 5px 15px;
    color: white;
}
QPushButton:hover {
    background: #0062A3;
}
QPushButton:pressed {
    background: #004F73;
}
QPushButton:disabled {
    background-color: #555555;
    color: #AAAAAA;
    border: 1px solid #3F3F46;
}
QCheckBox {
    spacing: 5px;
}
QProgressBar {
    border: 1px solid #3F3F46;
    text-align: center;
    background: #252526;
}
QProgressBar::chunk {
    background: #007ACC;
}
"""

DARK_STYLE_2 = """
QWidget {
    background-color: #2D2D30;
    color: #DCDCDC;
}
QTabWidget::pane {
    border: 1px solid #3F3F46;
    background: #2D2D30;
}
QHeaderView::section {
        background-color: #3F3F46;
        color: white;
        border: 1px solid #2D2D30;
        padding: 4px;
}
QTableCornerButton::section {
    background-color: #3F3F46;
    border: 1px solid #2D2D30;
}
QTabBar::tab {
    background: #252526;
    color: #DCDCDC;
    padding: 8px;
    border: 1px solid #3F3F46;
}
QTabBar::tab:selected {
    background: #3E3E42;
}
QLineEdit {
    background: #252526;
    border: 1px solid #3F3F46;
    padding: 3px;
}
QPushButton {
    background: #007ACC;
    border: none;
    padding: 5px 15px;
    color: white;
}
QPushButton:hover {
    background: #0062A3;
}
QPushButton:pressed {
    background: #004F73;
}
QPushButton:disabled {
    background-color: #555555;
    color: #AAAAAA;
    border: 1px solid #3F3F46;
}
QCheckBox {
    spacing: 5px;
}
QProgressBar {
    border: 1px solid #3F3F46;
    text-align: center;
    background: #252526;
}
QProgressBar::chunk {
    background: #007ACC;
}
"""

class QTextEditLogger:
    def __init__(self, text_edit):
        self.text_edit = text_edit

    def write(self, message):
        self.text_edit.append(f"[{datetime.now().strftime('%H:%M:%S')}] " + message.strip())  # Append text to QTextEdit

    def flush(self):  # Required for sys.stdout
        pass

class ConfigSyncWorker(QThread):
    log_signal = pyqtSignal(str)
    return_signal = pyqtSignal(list, list)

    def __init__(self, raw_path, data_path):
        super().__init__()
        self.raw_path = raw_path
        self.data_path = data_path

    def run(self):
        original_stdout = sys.stdout 
        sys.stdout = EmittingStream(self.log_signal)
        try:
            self.log_signal.emit("Starting LED detection...")
            vid_path, cfg_path = mky.configSync()
            self.return_signal.emit(vid_path, cfg_path)
            self.log_signal.emit("Finished LED detection\n")
        except Exception as e:
            self.log_signal.emit(f"Error during configSync: {str(e)}\n")
        finally:
            sys.stdout = original_stdout 

class EmittingStream:
    """ Redirects stdout to a PyQt signal. """
    def __init__(self, signal):
        self.signal = signal

    def write(self, text):
        if text.strip():
            self.signal.emit(str(text))

    def flush(self):
        pass  # Required for file-like objects


class RunSyncWorker(QThread):
    log_signal = pyqtSignal(str)
    # return_signal = pyqtSignal(list, list)

    def __init__(self, vid_path, cfg_path):
        super().__init__()
        self.vid_path = vid_path
        self.cfg_path = cfg_path

    def run(self):
        original_stdout = sys.stdout 
        sys.stdout = EmittingStream(self.log_signal)
        try:
            self.log_signal.emit("Starting vid sync...")
            mky.syncVid(self.vid_path, self.cfg_path)
            # self.return_signal.emit(vid_path, cfg_path)
            self.log_signal.emit("Finished vid sync\n")
        except Exception as e:
            self.log_signal.emit(f"Error during Sync: {str(e)}\n")
        finally:
            sys.stdout = original_stdout 

class ToColabWorker(QThread):
    log_signal = pyqtSignal(str)
    # return_signal = pyqtSignal(list, list)

    def __init__(self, vid_path):
        super().__init__()
        self.vid_path = vid_path

    def run(self):
        original_stdout = sys.stdout 
        sys.stdout = EmittingStream(self.log_signal)
        try:
            self.log_signal.emit("Moving videos to Colab...")
            mky.copyToGoogle(self.vid_path)
            # self.return_signal.emit(vid_path, cfg_path)
            self.log_signal.emit("Moved videos to Colab!\n")
        except Exception as e:
            self.log_signal.emit(f"Error during moving: {str(e)}\n")
        finally:
            sys.stdout = original_stdout 

class FromColabWorker(QThread):
    log_signal = pyqtSignal(str)
    # return_signal = pyqtSignal(list, list)

    def __init__(self, animal, date):
        super().__init__()
        self.animal = animal
        self.date = date

    def run(self):
        original_stdout = sys.stdout 
        sys.stdout = EmittingStream(self.log_signal)
        try:
            self.log_signal.emit("Fetching h5 from Colab...")
            mky.pickupFromGoogle(self.animal, self.date)
            # self.return_signal.emit(vid_path, cfg_path)
            self.log_signal.emit("Fetched h5 from Colab!\n")
        except Exception as e:
            self.log_signal.emit(f"Error during moving: {str(e)}\n")
        finally:
            sys.stdout = original_stdout

class SetupAniposeWorker(QThread):
    log_signal = pyqtSignal(str)
    # return_signal = pyqtSignal(list, list)

    def __init__(self, ani_base_path, vid_path):
        super().__init__()
        self.ani_base_path = ani_base_path
        self.vid_path = vid_path

    def run(self):
        original_stdout = sys.stdout 
        sys.stdout = EmittingStream(self.log_signal)
        try:
            self.log_signal.emit("Setting up anipose folder structure...")
            mky.setupAnipose(self.ani_base_path, self.vid_path)
            # self.return_signal.emit(vid_path, cfg_path)
            self.log_signal.emit("Set up anipose folder structure!\n")
        except Exception as e:
            self.log_signal.emit(f"Error during setting up: {str(e)}\n")
        finally:
            sys.stdout = original_stdout

class RunAniposeWorker(QThread):
    log_signal = pyqtSignal(str)
    # return_signal = pyqtSignal(list, list)

    def __init__(self, ani_base_path, run_combined):
        super().__init__()
        self.ani_base_path = ani_base_path
        self.run_combined = run_combined

    def run(self):
        original_stdout = sys.stdout 
        sys.stdout = EmittingStream(self.log_signal)
        try:
            self.log_signal.emit("Starting anipose... Don't click anything else nor exit!!")
            self.log_signal.emit("Check status in command lines")
            mky.runAnipose(self.ani_base_path, self.run_combined)
            # self.return_signal.emit(vid_path, cfg_path)
            self.log_signal.emit("Finished anipose!\n")
        except Exception as e:
            self.log_signal.emit(f"Error during moving: {str(e)}\n")
        finally:
            sys.stdout = original_stdout

class PipelineGUI(QMainWindow):
    def __init__(self):
        super().__init__()
        self.initUI()
        self.setupConnections()
        
    def initUI(self):
        self.setWindowTitle("Monkey Unity Pipeline v1.5")
        self.setGeometry(100, 100, 750, 630)
        self.setStyleSheet(DARK_STYLE)

        self.cam_header = ['Camera files \n(1 LR)','Camera files \n(2 LL)', 'Camera files (3 RR)', 'Camera files (4 RL)']
        
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        
        layout = QVBoxLayout()
        main_widget.setLayout(layout)
        
        # Tabs
        self.tabs = QTabWidget()
        self.tabs.currentChanged.connect(self.adjust_tab_height)
        layout.addWidget(self.tabs)
        
        # Setup Tab
        self.setup_tab = QWidget()
        self.filename_cam_head = ['x','x','x','x']
        self.createSetupTab()
        self.tabs.addTab(self.setup_tab, "Data Setup")
        
        # Sync Tab
        self.sync_tab = QWidget()
        self.createSyncTab()
        self.tabs.addTab(self.sync_tab, "Video Sync")
        
        # DLC Tab
        self.dlc_tab = QWidget()
        self.createDLCTab()
        self.tabs.addTab(self.dlc_tab, "DeepLabCut")
        
        # Anipose Tab
        self.anipose_tab = QWidget()
        self.createAniposeTab()
        self.tabs.addTab(self.anipose_tab, "Anipose")

        # Tool Tab
        self.tool_tab = QWidget()
        self.createToolTab()
        self.tabs.addTab(self.tool_tab, 'Tools')

        self.adjust_tab_height()
        
        # Log Output
        self.log_area = QTextEdit()
        self.log_area.setReadOnly(True)
        layout.addWidget(self.log_area)
        sys.stdout = QTextEditLogger(self.log_area)
        print('Test!')

        mky.list_skipped_file_name = ['x', '-']
    
    def adjust_tab_height(self):
        """Adjust height dynamically based on the current tab's content."""
        current_tab = self.tabs.currentWidget()
        if current_tab:
            self.tabs.setFixedHeight(current_tab.sizeHint().height() + self.tabs.tabBar().height())

    def createSetupTab(self):
        l = QVBoxLayout()
        self.data_setup_grp = QGroupBox("Data setup")
        layout = QFormLayout()

        self.raw_path = QLineEdit(mky.pm.PPATH_RAW)
        self.btn_path_refresh = QPushButton("Refresh")
        self.btn_browse_raw = QPushButton("Browse")
        self.btn_today = QPushButton("Today")
        self.btn_data_setup = QPushButton("Setup data folder")
        path_layout = QHBoxLayout()
        path_layout.addWidget(self.raw_path)
        path_layout.addWidget(self.btn_browse_raw)
        layout.addRow("RAW Data Path:", path_layout)
        path_layout = QHBoxLayout()
        path_layout.addWidget(self.btn_today)
        path_layout.addWidget(self.btn_path_refresh)
        layout.addRow(path_layout)
        
        self.animal_list = QLineEdit(",".join(mky.ANIMALS))
        layout.addRow("Animals:", self.animal_list)
        
        self.header_key = QLineEdit(mky.HEADER_KEY)
        layout.addRow("Header Key:", self.header_key)

        '''self.cam_offset = []
        offset_layout = QHBoxLayout()
        for i in range(4):
            sp = QSpinBox(self)
            sp.setRange(0, 9999)
            if i==0:
                sp.setValue(mky.CAM_OFFSETS[1])
            else:
                sp.setValue(mky.CAM_OFFSETS[i+1] + mky.CAM_OFFSETS[1])
            self.cam_offset.append(sp)
            offset_layout.addWidget(QLabel(f"CAM {i+1}"))
            offset_layout.addWidget(self.cam_offset[i])
            offset_layout.addStretch()'''
        # layout.addRow(offset_layout) 

        layout.addRow(self.btn_data_setup)

        self.data_setup_grp.setLayout(layout)
        l.addWidget(self.data_setup_grp)

        # Create a scrollable widget to display non-void rows
        self.non_void_scroll_area = QScrollArea()
        self.non_void_scroll_area.setWidgetResizable(True)
        
        self.non_void_container = QWidget()
        self.non_void_layout = QVBoxLayout()
        self.non_void_container.setLayout(self.non_void_layout)
        
        self.non_void_scroll_area.setWidget(self.non_void_container)
        l.addWidget(self.non_void_scroll_area)
        
        # Populate non-void rows
        self.populateNonVoidRows()

        self.setup_tab.setLayout(l)
        
    def createSyncTab(self):
        layout = QFormLayout()
        
        self.roi_table = QTableWidget()
        self.roi_table.setColumnCount(6)
        self.roi_table.setHorizontalHeaderLabels(["Cam", "X", "Y", "W", "H", 'LED'])
        self.populateROITable()
        self.roi_table.setSizeAdjustPolicy(QAbstractScrollArea.AdjustToContents)
        self.roi_table.setMaximumHeight(180)
        self.roi_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.roi_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addRow("Sync ROI config:", self.roi_table)

        self.sync_cam_chk = []
        self.sync_cam_btn_check = []
        self.sync_cam_btn_set = []
        self.sync_cam_cmb = []
        for j in range(2):
            cam_layout = QHBoxLayout()
            for i in range(1,3):
                chk = QCheckBox(f"CAM {i+2*j}")
                chk.setChecked(True)
                cam_layout.addWidget(chk)
                btn1 = QPushButton("Check ROI")
                btn2 = QPushButton("Set ROI")
                cmb = QComboBox()
                cmb.addItems(['Y', 'G'])
                cmb.setCurrentText(mky.LEDs[i+2*j]) 
                cam_layout.addWidget(btn1)
                cam_layout.addWidget(btn2)
                cam_layout.addWidget(cmb)
                cam_layout.addStretch()
                self.sync_cam_chk.append(chk)
                self.sync_cam_btn_check.append(btn1)
                self.sync_cam_btn_set.append(btn2)
                self.sync_cam_cmb.append(cmb)
            layout.addRow("",cam_layout)
        
        self.thres = QSpinBox()
        self.thres.setRange(100, 255)
        self.thres.setValue(mky.THRES)
        layout.addRow("Threshold:", self.thres)

        self.btn_sync_detect = QPushButton("Detect LED")
        self.btn_sync_run = QPushButton("Run sync")
        lt = QHBoxLayout()
        lt.addWidget(self.btn_sync_detect)
        lt.addWidget(self.btn_sync_run)
        layout.addRow(lt)
        
        self.sync_tab.setLayout(layout)

    def createDLCTab(self):
        layout = QVBoxLayout()
        
        self.dlc_mode = QButtonGroup()
        self.local_dlc = QRadioButton("Local DLC")
        self.colab_dlc = QRadioButton("Google Colab")
        self.dlc_mode.addButton(self.local_dlc)
        self.dlc_mode.addButton(self.colab_dlc)
        self.colab_dlc.setChecked(True)
        
        mode_layout = QHBoxLayout()
        mode_layout.addWidget(QLabel("Processing Mode:"))
        mode_layout.addWidget(self.local_dlc)
        mode_layout.addWidget(self.colab_dlc)
        layout.addLayout(mode_layout)
        
        # Local DLC Settings
        self.local_dlc_group = QGroupBox("Local DLC Settings")
        local_layout = QFormLayout()
        self.dlc_cfg_pathL = QLineEdit(mky.dlc_mdl_path['L'])
        self.btn_dlc_cfgL = QPushButton("Browse")
        self.dlc_cfg_pathR = QLineEdit(mky.dlc_mdl_path['R'])
        self.btn_dlc_cfgR = QPushButton("Browse")
        lt = QHBoxLayout()
        lt.addWidget(self.dlc_cfg_pathL)
        lt.addWidget(self.btn_dlc_cfgL)
        local_layout.addRow("Config Path Left:", lt)
        lt = QHBoxLayout()
        lt.addWidget(self.dlc_cfg_pathR)
        lt.addWidget(self.btn_dlc_cfgR)
        local_layout.addRow("Config Path Right:", lt)

        self.btn_run_dlc = QPushButton('今已知汝名，汝急速去--急急如律令!!')
        local_layout.addRow(self.btn_run_dlc)

        self.local_dlc_group.setLayout(local_layout)
        layout.addWidget(self.local_dlc_group)
        
        # Colab Settings
        self.colab_group = QGroupBox("Colab Settings")
        colab_layout = QFormLayout()

        self.colab_pathL = QLineEdit(mky.model_path_colab['L'])
        self.btn_colab_pathL = QPushButton("Browse")
        lt = QHBoxLayout()
        lt.addWidget(self.colab_pathL)
        lt.addWidget(self.btn_colab_pathL)
        colab_layout.addRow("Colab Path Left:", lt)

        self.colab_pathR = QLineEdit(mky.model_path_colab['R'])
        self.btn_colab_pathR = QPushButton("Browse")
        lt = QHBoxLayout()
        lt.addWidget(self.colab_pathR)
        lt.addWidget(self.btn_colab_pathR)
        colab_layout.addRow("Colab Path Right:", lt)

        self.btn_toColab = QPushButton("Move *.mp4 to Colab")
        self.btn_fromColab = QPushButton("Fetch *.h5 from Colab")
        lt = QHBoxLayout()
        lt.addWidget(self.btn_toColab)
        lt.addWidget(self.btn_fromColab)
        colab_layout.addRow(lt)

        self.colab_group.setLayout(colab_layout)
        self.colab_group.setEnabled(False)
        layout.addWidget(self.colab_group)
        
        self.dlc_tab.setLayout(layout)
        self.dlc_pane_stat_chg()

    def createAniposeTab(self):
        layout = QFormLayout()
        
        '''
        # this is discarded since currently we dont add points
        self.ref_table = QTableWidget()
        self.ref_table.setColumnCount(5)
        self.ref_table.setHorizontalHeaderLabels(["Cam", "Point", "X", "Y", "Scale"])
        self.populateRefTable()
        layout.addRow("Reference Points:", self.ref_table)
        
        self.scale_factor = QDoubleSpinBox()
        self.scale_factor.setValue(mky.SCALE_FACTOR)
        layout.addRow("Scale Factor:", self.scale_factor)'''

        self.chk_ani_label_combined = QCheckBox('Run `label-combined`')
        self.btn_run_ani = QPushButton('Run Anipose!')
        self.edt_anicfg = QLineEdit(mky.ani_cfg_mothercopy)
        self.btn_anicfg = QPushButton("Browse")
        self.btn_setup_ani = QPushButton('Setup anipose folder')
        lt = QHBoxLayout()
        lt.addWidget(self.edt_anicfg)
        lt.addWidget(self.btn_anicfg)
        layout.addRow(lt)
        layout.addRow(self.chk_ani_label_combined)
        lt = QHBoxLayout()
        lt.addWidget(self.btn_setup_ani)
        lt.addWidget(self.btn_run_ani)
        layout.addRow(lt)
        
        self.anipose_tab.setLayout(layout)

    def createToolTab(self):
        # l = QVBoxLayout()
        lt = QFormLayout()
        # grp = QGroupBox()
        self.edt_xlsx_path = QLineEdit()
        self.btn_xlsx_browse = QPushButton('Browse')
        ltt = QHBoxLayout()
        ltt.addWidget(self.edt_xlsx_path)
        ltt.addWidget(self.btn_xlsx_browse)
        lt.addRow("Exp note path:", ltt)
        self.btn_xlsx_fill = QPushButton('Fill file names')
        lt.addRow(self.btn_xlsx_fill)
        self.tool_tab.setLayout(lt)
    
    def setupConnections(self):
        self.raw_path.returnPressed.connect(self.update_raw_path)
        self.btn_browse_raw.clicked.connect(self.browseRawPath)
        self.btn_today.clicked.connect(self.setPathToday)
        self.btn_path_refresh.clicked.connect(self.pathRefresh)
        self.local_dlc.toggled.connect(self.dlc_pane_stat_chg)
        self.btn_data_setup.clicked.connect(self.dataSetup)
        self.btn_sync_detect.clicked.connect(self.btnDetect)
        self.btn_sync_run.clicked.connect(self.btnRunSync)
        self.btn_toColab.clicked.connect(self.toColab)
        self.btn_fromColab.clicked.connect(self.fromColab)
        self.btn_setup_ani.clicked.connect(self.setupAnipose)
        self.btn_run_ani.clicked.connect(self.runAnipose)
        for i in range(4):
            self.sync_cam_btn_set[i].clicked.connect(self.setROI)
            self.sync_cam_btn_check[i].clicked.connect(self.checkROI)
            self.sync_cam_cmb[i].currentIndexChanged.connect(self.cmbCamLED)
            # self.cam_offset[i].valueChanged.connect(self.camNo)
        self.btn_xlsx_fill.clicked.connect(self.fillExpNote)
        
    def update_raw_path(self):
        mem = mky.pm.PPATH_RAW
        mky.pm.PPATH_RAW = self.raw_path.text()
        try:
            self.populateNonVoidRows()
        except Exception as e:
            print(f'Error updating task panes {e}')
            print('Path **NOT** updated')
            mky.pm.PPATH_RAW = mem
            self.raw_path.setText(mky.pm.PPATH_RAW)
            return
        
        print(f"Updated PPATH_RAW: {mky.pm.PPATH_RAW}")
        animal, _ = mky._infoFromPath(mky.pm.PPATH_RAW)
        self.animal_list.setText(animal)

    def dlc_pane_stat_chg(self):
        b = self.local_dlc.isChecked()
        self.colab_group.setEnabled(not b)
        self.local_dlc_group.setEnabled(b)
        '''self.btn_colab_pathL.setEnabled(not b) # just to make it grey
        self.btn_colab_pathR.setEnabled(not b)
        self.btn_toColab.setEnabled(not b)
        self.btn_fromColab.setEnabled(not b)
        self.btn_dlc_cfgL.setEnabled(b) # just to make it grey
        self.btn_dlc_cfgR.setEnabled(b)
        self.btn_run_dlc.setEnabled(b)'''

    def browseRawPath(self):
        path = QFileDialog.getExistingDirectory(self, "Select RAW Data Directory", r'P:\projects\monkeys\Chronic_VLL\DATA_RAW\Pici')
        if path:
            self.raw_path.setText(path)
            self.update_raw_path()
    
    def setPathToday(self):
        try:
            path = os.path.join(r'P:\projects\monkeys\Chronic_VLL\DATA_RAW\Pici', datetime.today().strftime('%Y'), datetime.today().strftime('%m'), datetime.today().strftime('%Y%m%d'))
            self.raw_path.setText(path)
            self.update_raw_path()
        except Exception as e:
            print(e)

    def pathRefresh(self):
        '''animal, _ = mky._infoFromPath(mky.pm.PPATH_RAW)
        self.animal_list.setText(animal)'''
        self.update_raw_path()
    
    def dataSetup(self):
        # self.df = mky.readExpNote(PPATH_RAW)
        if self.btn_data_setup.text() != 'Confirm by clicking again':
            self.btn_data_setup.setText('Confirm by clicking again')
            print(f'Will setup for {mky.pm.PPATH_RAW}!')
            QTimer.singleShot(3000, lambda: self.btn_data_setup.setText('Setup data folder'))
            # dont say its stupid!
        else:
            self.btn_data_setup.setText('Setup data folder')
            if os.path.exists(mky.pm.data_path):
                print(f'Folder already exists for {os.path.basename(mky.pm.data_path)}')
                if not os.path.exists(os.path.join(mky.pm.data_path, f'{os.path.basename(mky.pm.PPATH_RAW)}.lnk')):
                    mky.two_way_shortcuts(mky.pm.data_path, mky.pm.PPATH_RAW) # still setup folder
            else:
                mky.dataSetup()       
                mky.two_way_shortcuts(mky.pm.data_path, mky.pm.PPATH_RAW)
                print('Folder has been setup')


    def populateNonVoidRows(self):
        """Populate non-void rows inside the scroll area."""
        df = mky.readExpNote(mky.pm.PPATH_RAW, header=['Experiment Number', 'Experiment', 'Task', 'VOID',
                                                    'Camera files \n(1 LR)','Camera files \n(2 LL)', 'Camera files (3 RR)', 'Camera files (4 RL)'
                                                    ])     
        animal, date = mky.pm.animal, mky.pm.date
            
        # Remove existing widgets
        for i in reversed(range(self.non_void_layout.count())):
            widget = self.non_void_layout.itemAt(i).widget()
            if widget:
                widget.deleteLater()

        flag_head = True
        for _, row in df.iterrows():
            if row["VOID"] == "T":  # Skip void rows
                continue

            # Create a group box for each row
            exp_group = QGroupBox(f"{row['Experiment']}-{row['Task']}")
            lt = QGridLayout()

            lt.addWidget(QLabel(f"Experiment: {row['Experiment']}"),0,0)
            lt.addWidget(QLabel(f"Task: {row['Task']}"),0,2)
            # lt.addWidget(QLabel(f"Video # From: {row['Video # From']}"),1,0)
            # lt.addWidget(QLabel(f"Video # To: {row['Video # To']}"),1,2)
  
            y = '✅'
            n = '❌'
            wng = '⚠️'
            try:
                for i in range(4):
                    fname = f'C{int(row[self.cam_header[i]]):04}.{mky.vid_type}'
                    if os.path.exists(os.path.join(mky.pm.PPATH_RAW,f'cam{i+1}',fname)):
                        c = f'{fname}'
                    else: 
                        c = f'{fname} {wng}'
                        print(f'{wng} Video not found for {row["Experiment"]}-{row["Task"]}!')
                    lt.addWidget(QLabel(f'Cam {i+1}: {c}'), 1, i)
                    if flag_head:
                        self.filename_cam_head[i] = (f'C{int(row[self.cam_header[i]]):04}.{mky.vid_type}')  # just used for ROI check
                        '''if flag_head:
                        fname = f'C{int(row[self.cam_header[i]]):04}.{mky.vid_type}'
                        # if not os.path.exists(os.path.join(mky.pm.PPATH_RAW,f'cam{i+1}',fname)):
                        #     raise FileNotFoundError(f'Video file {fname} not found for {row['Experiment']}-{row["Task"]}')
                        self.filename_cam_head[i] = (f'{fname}') if os.path.exists(os.path.join(mky.pm.PPATH_RAW,f'cam{i+1}',fname)) else (f'{fname}{n}')'''
                flag_head = False
                # print(self.filename_cam_head)
            except Exception as e:
                print(f"Error generating file name labels for {row['Experiment']}-{row['Task']}, check exp notes for cam files 1~4. {e}")

            p = os.path.join(mky.pm.data_path, 'SynchronizedVideos', f"{date}-{animal}-{row['Experiment']}-{row['Task']}")
            caps = ['Detected', 'Sync-ed', 'DLC processed','CSV out']
            stat = []
            stat.append(y if os.path.exists(os.path.join(p, '.skipDet')) else n)
            stat.append(y if os.path.exists(os.path.join(p, '.skipSync')) else n)
            c = sum([os.path.exists(os.path.join(p, 'L', '.skipDLC')), os.path.exists(os.path.join(p, 'R', '.skipDLC'))])
            if c == 2:
                stat.append(y)
            elif c == 1:
                stat.append('🤔')
            else:
                stat.append(n)
            if os.path.exists(os.path.join(mky.pm.data_path, 'anipose', f"{date}-{animal}-{row['Experiment']}-{row['Task']}",
                                                         "pose-3d",
                                                         f"{date}-{animal}-{row['Experiment']}-{row['Task']}.csv")):
                stat.append(y+'🎊')
            else:
                stat.append(n)
            for i in range(4):
                lt.addWidget(QLabel(caps[i] + stat[i]), 2, i)
            
            exp_group.setLayout(lt)
            self.non_void_layout.addWidget(exp_group)

        self.non_void_layout.addStretch()
    
    def fillExpNote(self):
        path = self.edt_xlsx_path.text()
        wb = openpyxl.load_workbook(path)
        wb.save(os.path.join(os.path.dirname(path), f'originalCopy_{os.path.basename(path)}'))
        if wb:
            ws = wb.active
            hr = None
            for row in ws.iter_rows():
                if any(cell.value == self.header_key.text() for cell in row):
                    hr = row[0].row
                break
            if not hr:
                print(f'No header found with key {self.edt_xlsx_path.text()}')
                return
            
            camnum_idx = {}
            for c in ws[hr]:
                if c.value in self.cam_header:
                    camnum_idx[c.value] = c.column
            
            for r in ws.iter_rows(min_row = hr + 1, max_row= ws.max_row):
                for cname, cidx in camnum_idx.items():
                    cell = r[cidx - 1]
                    v = cell.value
                    if v in mky.list_skipped_file_name: # sign for actually no file
                        continue
                    if v is None:
                        row = cell.row
                        while row > hr:
                            row -= 1
                            acell = ws.cell(row = row, column = cidx)
                            if acell.value is not None:
                                ahcell = ws.cell(row = row, column = 0)
                                hcell = ws.cell(row = cell.row, column = 0)
                                r[cidx - 1].value = acell.value + (hcell.value - ahcell.value)
                                break
                    # r[cidx - 1].value = 'x'
                print(r)
            
        wb.save(os.path.join(os.path.dirname(path), f'{os.path.basename(path)}'))
        wb.close()

    '''def camNo(self):
        try: 
            if self.sender() in self.cam_offset:
                idx = self.cam_offset.index(self.sender())
                if idx != 0:
                    idx = self.cam_offset.index(self.sender())
                    mky.CAM_OFFSETS[idx+1] = int(self.cam_offset[idx].value()-self.cam_offset[0].value())
                    self.cam_offset[idx].setValue(mky.CAM_OFFSETS[1] + mky.CAM_OFFSETS[idx+1])
                else:
                    for i in range(1,4):
                        mky.CAM_OFFSETS[i+1] = int(self.cam_offset[i].value() - self.cam_offset[0].value())
                    mky.CAM_OFFSETS[1] =  int(self.cam_offset[0].value())
            else:
                print('Unrecognized SpinBox!')
        except Exception as e:
            print(f'Error in self.camNo: {e}')'''

    def populateROITable(self):
        self.roi_table.setRowCount(0)#len(mky.ROIs))
        for i, (cam, vals) in enumerate(mky.ROIs.items()):
            self.roi_table.insertRow(i)
            self.roi_table.setItem(i, 0, QTableWidgetItem(str(cam)))
            for j in range(4):
                self.roi_table.setItem(i, j+1, QTableWidgetItem(str(vals[j])))
            self.roi_table.setItem(i, 5, QTableWidgetItem(str(mky.LEDs[i+1])))
    
    def populateRefTable(self):
        self.ref_table.setRowCount(len(mky.Ref)*3)
        row = 0
        for cam, points in enumerate(mky.Ref):
            for point, coords in points.items():
                self.ref_table.insertRow(row)
                self.ref_table.setItem(row, 0, QTableWidgetItem(str(cam+1)))
                self.ref_table.setItem(row, 1, QTableWidgetItem(point))
                for i in range(3):
                    self.ref_table.setItem(row, 2+i, QTableWidgetItem(str(coords[i])))
                row += 1
    
    def checkROI(self):
        if self.sender() in self.sync_cam_btn_check:
            try:
                idx = self.sync_cam_btn_check.index(self.sender())
            except Exception:
                print('Alien invasion detected when trying to locate check button')   

            if self.filename_cam_head[idx] in mky.list_skipped_file_name:
                print('Alien invasion detected when trying to determine file name')
                return
            p = os.path.join(mky.pm.PPATH_RAW, f'cam{idx+1}', self.filename_cam_head[idx])
            r = mky.ROIs[idx+1]
            print(f'Trying to check {p}, {r}, frame {200}!')
            ROIConfig.show_saved_rois(p, r, 200)
                
        # ROIConfig.

    def setROI(self):
        try:
            if self.sender() in self.sync_cam_btn_set:
                idx = self.sync_cam_btn_set.index(self.sender())
                # print(idx)
                path = os.path.join(mky.pm.PPATH_RAW, f'cam{idx+1}', self.filename_cam_head[idx])
                frame = 200
                ROI = ROIConfig.draw_roi(path, frame)
                if ROI is None:
                    # print('No ROI is selected')
                    return
                mky.ROIs[idx+1] = ROI
                print(f'Updated ROI: cam {idx} path {os.path.basename(path)} at frame {frame+1}, ROI {ROI}')
                self.populateROITable()
            else:
                print('Unidentified button!')
        except Exception as e:
            print(f'Error in self.setROI: {e}')
    
    def btnDetect(self):
        self.update_raw_path()
        mky.dataSetup()

        # Threaded execution of configSync
        self.config_sync_worker = ConfigSyncWorker(mky.pm.PPATH_RAW, mky.pm.data_path)
        self.config_sync_worker.log_signal.connect(self.log_area.append)
        self.config_sync_worker.return_signal.connect(self.handleConfigSyncResult)
        self.config_sync_worker.finished.connect(self.btnDetSyncDone)
        self.config_sync_worker.start()
        self.btn_sync_run.setEnabled(False)
        self.btn_sync_detect.setEnabled(False)
        self.btn_sync_run.setText('Detecting...')
        self.btn_sync_detect.setText('Detecting...')

    def handleConfigSyncResult(self, vid_path, cfg_path):
        mky.pm.vid_path = vid_path
        mky.pm.cfg_path = cfg_path
        self.log_area.append("configSync results received. Ready for the next step.\n")
    
    def btnDetSyncDone(self):
        self.config_sync_worker.deleteLater()
        self.syncDone()
    
    def cmbCamLED(self):
        try:
            if self.sender() in self.sync_cam_cmb:
                idx = self.sync_cam_cmb.index(self.sender())
                # print(idx)
                LED = self.sync_cam_cmb[idx].currentText()
                if LED in ['Y', 'G']:
                    mky.LEDs[idx+1] = LED
                else:
                    print('sth strange happened in LED setting')
                print(f'Updated LED: cam {idx+1} LED {LED}')
                self.populateROITable()
            else:
                print('Unidentified button!')
        except Exception as e:
            print(f'Error in self.cmbCamLED: {e}')

    def btnRunSync(self):
        if not mky.pm.vid_path or not mky.pm.cfg_path:
            print('Plz wait for configSync() and run again later')
            self.btnDetect()
            return
        self.run_sync_worker = RunSyncWorker(mky.pm.vid_path, mky.pm.cfg_path)
        self.run_sync_worker.log_signal.connect(self.log_area.append)
        self.run_sync_worker.finished.connect(self.btnRunSyncDone)
        self.run_sync_worker.start()
        self.btn_sync_run.setEnabled(False)
        self.btn_sync_detect.setEnabled(False)
        self.btn_sync_run.setText('Synchronizing...')
        self.btn_sync_detect.setText('Synchronizing...')
    
    def btnRunSyncDone(self):
        self.run_sync_worker.deleteLater()
        self.syncDone()
    
    def syncDone(self):
        self.btn_sync_run.setEnabled(True)
        self.btn_sync_detect.setEnabled(True)
        self.btn_sync_run.setText('Run sync')
        self.btn_sync_detect.setText('Detect LED')

    def toColab(self):
        self.update_raw_path()
        if mky.pm.vid_path is None:
            print('Plz wait for configSync() and run again later')
            self.btnDetect()
            return
            '''self.btnDetect()
            print('Plz wait for configSync() and run again later')
            return'''
        self.to_colab_worker = ToColabWorker(mky.pm.vid_path)
        self.to_colab_worker.log_signal.connect(self.log_area.append)
        self.to_colab_worker.finished.connect(self.to_colab_worker.deleteLater)
        self.to_colab_worker.start()

    def fromColab(self):
        self.from_colab_worker = FromColabWorker(mky.pm.animal, mky.pm.date)
        self.from_colab_worker.log_signal.connect(self.log_area.append)
        self.from_colab_worker.finished.connect(self.from_colab_worker.deleteLater)
        self.from_colab_worker.start()

    def setupAnipose(self):
        if mky.pm.vid_path is None:
            print("ValueError('data_path not configured!!') Plz Run detect LED first")
            self.btnDetect()
            return
        self.setup_anipose_worker = SetupAniposeWorker(mky.pm.ani_base_path, mky.pm.vid_path)
        self.setup_anipose_worker.log_signal.connect(self.log_area.append)
        self.setup_anipose_worker.finished.connect(self.setup_anipose_worker.deleteLater)
        self.setup_anipose_worker.start()

    def runAnipose(self):
        if not os.path.exists(os.path.join(mky.pm.ani_base_path, mky.ani_cfg_mothercopy)):
            print("RuntimeError('anipose folder not set up yet!')")
            return
        self.run_anipose_worker = RunAniposeWorker(mky.pm.ani_base_path, self.chk_ani_label_combined.isChecked())
        self.run_anipose_worker.log_signal.connect(self.log_area.append)
        self.run_anipose_worker.finished.connect(self.run_anipose_worker.deleteLater)
        self.run_anipose_worker.start()

        
    def logMessage(self, message):
        self.log_area.append(f"[{datetime.now().strftime('%H:%M:%S')}] {message}")
    
    def fullAuto(self):
        try:
            # Update parameters from GUI
            mky.pm.PPATH_RAW = self.raw_path.text()
            mky.ANIMALS = [a.strip() for a in self.animal_list.text().split(",")]
            
            # Run pipeline steps
            self.logMessage("Starting pipeline...")
            
            # Data Setup
            data_path = mky.dataSetup(mky.pm.PPATH_RAW)
            
            # Sync
            vid_path, cfg_path = mky.configSync(mky.pm.PPATH_RAW, data_path)
            
            # DLC Branch
            if self.local_dlc.isChecked():
                mky.runDLC(vid_path)
            else:
                kids = mky.copyToGoogle(vid_path)
                # Wait for Colab processing...
            
            # Anipose
            mky.setupAnipose(os.path.join(data_path, 'anipose'), 
                                vid_path, mky.Ref, mky.add_ref)
            mky.runAnipose()
            
            self.logMessage("Pipeline completed successfully!")
            
        except Exception as e:
            self.logMessage(f"Error: {str(e)}")
            QMessageBox.critical(self, "Error", str(e))

if __name__ == '__main__':
    app = QApplication(sys.argv)
    app.setFont(QFont("Segoe UI", 10))
    window = PipelineGUI()
    window.show()
    try:
        sys.exit(app.exec_())
    finally:
        with open(f"Log\GUI_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt", "w", encoding="utf-8") as f:
            f.write(window.log_area.toPlainText())
