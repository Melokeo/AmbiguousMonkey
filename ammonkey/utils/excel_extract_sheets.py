from datetime import datetime
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
import shutil

def copy_excel_tabs_to_files(
    excel_path: str | Path, 
    output_dir: str | Path = ".",
    copy_notice: str = "READ-ONLY: THIS IS A COPY"
) -> None:
    """copy each tab preserving formatting with datetime stamp"""
    
    excel_path = Path(excel_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(exist_ok=True)
    
    wb = load_workbook(excel_path)
    
    for sheet_name in wb.sheetnames:
        print(f"  copying sheet: {sheet_name}")
        
        source_ws = wb[sheet_name]
        new_wb = Workbook()
        new_ws = new_wb.active
        new_ws.title = sheet_name
        
        # copy all cells with formatting
        for row in source_ws.iter_rows():
            # check if entire row is empty
            if all(cell.value is None or cell.value == "" for cell in row):
                continue
                
            for cell in row:
                new_cell = new_ws.cell(row=cell.row, column=cell.column)
                new_cell.value = cell.value
                if cell.has_style:
                    new_cell.font = cell.font.copy()
                    new_cell.border = cell.border.copy()
                    new_cell.fill = cell.fill.copy()
                    new_cell.number_format = cell.number_format
                    new_cell.alignment = cell.alignment.copy()
        
        # add copy notice with timestamp
        dt = datetime.now().strftime('%Y-%m-%d %H:%M')
        last_row = new_ws.max_row + 2
        cell = new_ws.cell(row=last_row, column=1, value=f"{copy_notice} @ {dt}")
        cell.font = Font(bold=True)
        
        # save file
        output_file = output_dir / f"{sheet_name}.xlsx"
        new_wb.save(output_file)

def dispatch_files(source_dir: str | Path, year_dir: str | Path) -> None:
    """dispatch yyyymmdd files to yyyy/mm/ structure"""
    
    source_dir = Path(source_dir)
    year_dir = Path(year_dir)
    base_year = year_dir.name
    
    for file_path in source_dir.glob("*.xlsx"):
        filename = file_path.stem
        
        # check if filename matches yyyymmdd pattern
        if len(filename) == 8 and filename.isdigit():
            year = filename[:4]
            month = filename[4:6]
            
            if year == base_year:
                # create target directory
                target_dir = year_dir / month / filename
                if not target_dir.exists():
                    # raise FileNotFoundError(str(target_dir))
                    print(f'FileNotFound: {target_dir}. Did you map P: drive?')
                    continue
                target_file = target_dir / f"RISO_{filename}.xlsx"
                
                # handle duplicates
                if target_file.exists():
                    # target_file = target_dir / f"FUSILLO_{filename}_from_cloud.xlsx"
                    print(f'Skipped {target_file}: exists.')
                    continue
                
                shutil.move(str(file_path), str(target_file))
                print(f"moved {filename}.xlsx -> {target_file}")